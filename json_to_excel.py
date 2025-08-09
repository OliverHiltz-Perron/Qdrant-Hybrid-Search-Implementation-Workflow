#!/usr/bin/env python3
"""Convert JSON extraction results to Excel with one sheet per category."""

import json
import pandas as pd
from pathlib import Path
import argparse
from datetime import datetime
import logging
from openpyxl.styles import Alignment

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_json_results(json_path):
    """Load JSON extraction results with encoding detection."""
    # Try different encodings
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'windows-1252', 'cp1252']
    
    for encoding in encodings:
        try:
            with open(json_path, 'r', encoding=encoding) as f:
                data = json.load(f)
            logger.info(f"Successfully loaded JSON with {encoding} encoding")
            break
        except UnicodeDecodeError:
            continue
        except json.JSONDecodeError as e:
            # If JSON is invalid, try to clean it
            logger.warning(f"JSON decode error with {encoding}: {e}")
            try:
                with open(json_path, 'r', encoding=encoding, errors='replace') as f:
                    content = f.read()
                # Replace problematic characters
                content = content.replace('\x92', "'")  # Smart quote
                content = content.replace('\x93', '"')  # Smart quote
                content = content.replace('\x94', '"')  # Smart quote
                content = content.replace('\x96', '-')  # En dash
                content = content.replace('\x97', '--') # Em dash
                data = json.loads(content)
                logger.info(f"Successfully loaded JSON after cleaning with {encoding} encoding")
                break
            except:
                continue
    else:
        # If all encodings fail, try with error handling
        try:
            with open(json_path, 'r', encoding='utf-8', errors='replace') as f:
                data = json.load(f)
            logger.warning("Loaded JSON with replacement characters for invalid bytes")
        except Exception as e:
            logger.error(f"Failed to load JSON file: {e}")
            raise
    
    # Handle both single extraction and batch results
    if isinstance(data, list):
        # Batch results
        return data
    else:
        # Single result
        return [data]


def create_programs_sheet(results):
    """Create sheet for programs waiting to add."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        programs_data = result.get('extracted_data', {}).get('programs_waiting_to_add', {})
        
        if programs_data.get('programs'):
            for program in programs_data.get('programs', []):
                rows.append({
                    'State': state,
                    'Program Name': program.get('name', ''),
                    'Status': program.get('status', ''),
                    'Timeline': program.get('timeline', ''),
                    'Supporting Quote': program.get('quote', ''),
                    'Total Programs': programs_data.get('total_count', 0)
                })
        else:
            # If no programs, still add a row for the state
            rows.append({
                'State': state,
                'Program Name': 'None identified',
                'Status': '',
                'Timeline': '',
                'Supporting Quote': '',
                'Total Programs': 0
            })
    
    return pd.DataFrame(rows)


def create_populations_sheet(results):
    """Create sheet for target populations."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        pop_data = result.get('extracted_data', {}).get('target_populations', {})
        
        if pop_data.get('populations'):
            for pop in pop_data.get('populations', []):
                rows.append({
                    'State': state,
                    'Population Group': pop.get('group', ''),
                    'Criteria': pop.get('criteria', ''),
                    'Supporting Quote': pop.get('quote', ''),
                    'Primary Population': pop_data.get('primary_population', '')
                })
        else:
            rows.append({
                'State': state,
                'Population Group': 'None identified',
                'Criteria': '',
                'Supporting Quote': '',
                'Primary Population': ''
            })
    
    return pd.DataFrame(rows)


def create_eligibility_sheet(results):
    """Create sheet for eligibility determination."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        elig_data = result.get('extracted_data', {}).get('eligibility_determination', {})
        
        quotes = elig_data.get('quotes', [])
        screening_tools = elig_data.get('screening_tools', [])
        
        if quotes:
            # Create a row for each quote
            for i, quote in enumerate(quotes):
                rows.append({
                    'State': state,
                    'Determination Method': elig_data.get('determination_method', '') if i == 0 else '',
                    'Screening Tools': ', '.join(screening_tools) if i == 0 else '',
                    'Decision Maker': elig_data.get('decision_maker', '') if i == 0 else '',
                    'Supporting Quote': quote,
                    'Quote Number': f"{i+1} of {len(quotes)}"
                })
        else:
            # Single row if no quotes
            rows.append({
                'State': state,
                'Determination Method': elig_data.get('determination_method', ''),
                'Screening Tools': ', '.join(screening_tools),
                'Decision Maker': elig_data.get('decision_maker', ''),
                'Supporting Quote': '',
                'Quote Number': ''
            })
    
    return pd.DataFrame(rows)


def create_outcomes_sheet(results):
    """Create sheet for effectiveness outcomes."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        outcomes_data = result.get('extracted_data', {}).get('effectiveness_outcomes', {})
        
        has_goals = outcomes_data.get('has_outcome_goals', False)
        
        if outcomes_data.get('outcome_goals'):
            for goal in outcomes_data.get('outcome_goals', []):
                rows.append({
                    'State': state,
                    'Has Outcome Goals': 'Yes' if has_goals else 'No',
                    'Goal': goal.get('goal', ''),
                    'Metric': goal.get('metric', ''),
                    'Target': goal.get('target', ''),
                    'Supporting Quote': goal.get('quote', '')
                })
        else:
            rows.append({
                'State': state,
                'Has Outcome Goals': 'Yes' if has_goals else 'No',
                'Goal': 'None specified',
                'Metric': '',
                'Target': '',
                'Supporting Quote': ''
            })
    
    return pd.DataFrame(rows)


def create_monitoring_sheet(results):
    """Create sheet for monitoring and accountability."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        mon_data = result.get('extracted_data', {}).get('monitoring_accountability', {})
        
        quotes = mon_data.get('quotes', [])
        
        if quotes:
            # Create a row for each quote
            for i, quote in enumerate(quotes):
                rows.append({
                    'State': state,
                    'Has Monitoring System': 'Yes' if mon_data.get('has_monitoring_system', False) else 'No',
                    'CQI Mentioned': 'Yes' if mon_data.get('cqi_mentioned', False) else 'No',
                    'Fidelity Monitoring': 'Yes' if mon_data.get('fidelity_monitoring', False) else 'No',
                    'Monitoring Description': mon_data.get('monitoring_description', '') if i == 0 else '',
                    'Supporting Quote': quote,
                    'Quote Number': f"{i+1} of {len(quotes)}"
                })
        else:
            rows.append({
                'State': state,
                'Has Monitoring System': 'Yes' if mon_data.get('has_monitoring_system', False) else 'No',
                'CQI Mentioned': 'Yes' if mon_data.get('cqi_mentioned', False) else 'No',
                'Fidelity Monitoring': 'Yes' if mon_data.get('fidelity_monitoring', False) else 'No',
                'Monitoring Description': mon_data.get('monitoring_description', ''),
                'Supporting Quote': '',
                'Quote Number': ''
            })
    
    return pd.DataFrame(rows)


def create_workforce_sheet(results):
    """Create sheet for workforce support."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        work_data = result.get('extracted_data', {}).get('workforce_support', {})
        
        workforce = work_data.get('workforce_training', {})
        credentialing = work_data.get('credentialing', {})
        
        # Get all quotes
        training_quotes = workforce.get('quotes', [])
        cred_quotes = credentialing.get('quotes', [])
        
        # Determine max quotes
        max_quotes = max(len(training_quotes), len(cred_quotes), 1)
        
        for i in range(max_quotes):
            rows.append({
                'State': state,
                'Has Training Plan': 'Yes' if workforce.get('has_training_plan', False) else 'No',
                'Training Description': workforce.get('training_description', '') if i == 0 else '',
                'Training Quote': training_quotes[i] if i < len(training_quotes) else '',
                'Has Credentialing Requirements': 'Yes' if credentialing.get('has_requirements', False) else 'No',
                'Requirements': ', '.join(credentialing.get('requirements', [])) if i == 0 else '',
                'Credentialing Quote': cred_quotes[i] if i < len(cred_quotes) else '',
                'Quote Number': f"{i+1} of {max_quotes}" if max_quotes > 1 else ''
            })
    
    return pd.DataFrame(rows)


def create_funding_sheet(results):
    """Create sheet for funding sources."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        fund_data = result.get('extracted_data', {}).get('funding_sources', {})
        
        if fund_data.get('funding_sources'):
            for source in fund_data.get('funding_sources', []):
                rows.append({
                    'State': state,
                    'Funding Source': source.get('source', ''),
                    'Supporting Quote': source.get('quote', ''),
                    'Has Multiple Sources': 'Yes' if fund_data.get('has_multiple_sources', False) else 'No'
                })
        else:
            rows.append({
                'State': state,
                'Funding Source': 'None identified',
                'Supporting Quote': '',
                'Has Multiple Sources': 'No'
            })
    
    return pd.DataFrame(rows)


def create_trauma_sheet(results):
    """Create sheet for trauma-informed delivery."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        trauma_data = result.get('extracted_data', {}).get('trauma_informed_delivery', {})
        
        quotes = trauma_data.get('evidence_quotes', [])
        uses_trauma = trauma_data.get('uses_trauma_informed_approach', False)
        
        if quotes:
            # Create a row for each quote
            for i, quote in enumerate(quotes):
                rows.append({
                    'State': state,
                    'Uses Trauma-Informed Approach': 'Yes' if uses_trauma else 'No',
                    'Evidence Quote': quote,
                    'Quote Number': f"{i+1} of {len(quotes)}"
                })
        else:
            rows.append({
                'State': state,
                'Uses Trauma-Informed Approach': 'Yes' if uses_trauma else 'No',
                'Evidence Quote': '',
                'Quote Number': ''
            })
    
    return pd.DataFrame(rows)


def create_equity_sheet(results):
    """Create sheet for equity and disparity reduction."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        equity_data = result.get('extracted_data', {}).get('equity_disparity_reduction', {})
        
        quotes = equity_data.get('quotes', [])
        
        if quotes:
            # Create a row for each quote
            for i, quote in enumerate(quotes):
                rows.append({
                    'State': state,
                    'Addresses Equity': 'Yes' if equity_data.get('addresses_equity', False) else 'No',
                    'Equity Approach': equity_data.get('equity_approach', '') if i == 0 else '',
                    'Targeted Groups': ', '.join(equity_data.get('targeted_groups', [])) if i == 0 else '',
                    'Supporting Quote': quote,
                    'Quote Number': f"{i+1} of {len(quotes)}"
                })
        else:
            rows.append({
                'State': state,
                'Addresses Equity': 'Yes' if equity_data.get('addresses_equity', False) else 'No',
                'Equity Approach': equity_data.get('equity_approach', ''),
                'Targeted Groups': ', '.join(equity_data.get('targeted_groups', [])),
                'Supporting Quote': '',
                'Quote Number': ''
            })
    
    return pd.DataFrame(rows)


def create_structural_sheet(results):
    """Create sheet for structural determinants."""
    rows = []
    for result in results:
        state = result.get('metadata', {}).get('state', 'Unknown')
        struct_data = result.get('extracted_data', {}).get('structural_determinants', {})
        
        quotes = struct_data.get('quotes', [])
        
        if quotes:
            # Create a row for each quote
            for i, quote in enumerate(quotes):
                rows.append({
                    'State': state,
                    'Addresses Structural Determinants': 'Yes' if struct_data.get('addresses_structural_determinants', False) else 'No',
                    'Support Types': ', '.join(struct_data.get('support_types', [])) if i == 0 else '',
                    'Supporting Quote': quote,
                    'Quote Number': f"{i+1} of {len(quotes)}"
                })
        else:
            rows.append({
                'State': state,
                'Addresses Structural Determinants': 'Yes' if struct_data.get('addresses_structural_determinants', False) else 'No',
                'Support Types': ', '.join(struct_data.get('support_types', [])),
                'Supporting Quote': '',
                'Quote Number': ''
            })
    
    return pd.DataFrame(rows)


def create_summary_sheet(results):
    """Create summary sheet with overview of all states."""
    rows = []
    for result in results:
        if not result.get('success', True):
            continue
            
        state = result.get('metadata', {}).get('state', 'Unknown')
        extracted = result.get('extracted_data', {})
        
        row = {
            'State': state,
            'Document': result.get('metadata', {}).get('filename', ''),
            'Extraction Date': datetime.now().strftime('%Y-%m-%d'),
            'Programs Waiting': extracted.get('programs_waiting_to_add', {}).get('total_count', 0),
            'Has Outcome Goals': 'Yes' if extracted.get('effectiveness_outcomes', {}).get('has_outcome_goals', False) else 'No',
            'Has Monitoring System': 'Yes' if extracted.get('monitoring_accountability', {}).get('has_monitoring_system', False) else 'No',
            'Has Workforce Training': 'Yes' if extracted.get('workforce_support', {}).get('workforce_training', {}).get('has_training_plan', False) else 'No',
            'Multiple Funding Sources': 'Yes' if extracted.get('funding_sources', {}).get('has_multiple_sources', False) else 'No',
            'Trauma-Informed': 'Yes' if extracted.get('trauma_informed_delivery', {}).get('uses_trauma_informed_approach', False) else 'No',
            'Addresses Equity': 'Yes' if extracted.get('equity_disparity_reduction', {}).get('addresses_equity', False) else 'No',
            'Addresses Structural': 'Yes' if extracted.get('structural_determinants', {}).get('addresses_structural_determinants', False) else 'No'
        }
        rows.append(row)
    
    return pd.DataFrame(rows)


def main():
    parser = argparse.ArgumentParser(description='Convert JSON extraction results to Excel')
    parser.add_argument('json_file', help='Path to JSON extraction results')
    parser.add_argument('-o', '--output', help='Output Excel file path', 
                       default='extraction_results.xlsx')
    
    args = parser.parse_args()
    
    # Load JSON results
    logger.info(f"Loading JSON from {args.json_file}")
    results = load_json_results(args.json_file)
    
    # Filter for successful extractions
    successful_results = [r for r in results if r.get('success', True)]
    logger.info(f"Processing {len(successful_results)} successful extractions")
    
    # Create Excel writer
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Create summary sheet first
        logger.info("Creating summary sheet")
        summary_df = create_summary_sheet(successful_results)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create individual category sheets
        sheet_creators = {
            '1_Programs_Waiting': create_programs_sheet,
            '2_Target_Populations': create_populations_sheet,
            '3_Eligibility': create_eligibility_sheet,
            '4_Effectiveness': create_outcomes_sheet,
            '5_Monitoring': create_monitoring_sheet,
            '6_Workforce': create_workforce_sheet,
            '7_Funding': create_funding_sheet,
            '8_Trauma_Informed': create_trauma_sheet,
            '9_Equity': create_equity_sheet,
            '10_Structural': create_structural_sheet
        }
        
        for sheet_name, creator_func in sheet_creators.items():
            logger.info(f"Creating sheet: {sheet_name}")
            df = creator_func(successful_results)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Auto-adjust column widths after writing all sheets
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            # Check if it's a quote column
                            if 'quote' in str(worksheet.cell(1, cell.column).value).lower():
                                max_length = min(50, max(max_length, len(str(cell.value))))
                            else:
                                max_length = min(30, max(max_length, len(str(cell.value))))
                    except:
                        pass
                
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add text wrapping for quote columns
                if 'quote' in str(worksheet.cell(1, column[0].column).value).lower():
                    for cell in column:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    logger.info(f"Excel file saved to: {output_path}")
    print(f"\nExcel file created successfully: {output_path}")
    print(f"Contains {len(sheet_creators) + 1} sheets (Summary + 10 categories)")
    print("\nQuotes are now displayed on separate rows for better readability.")


if __name__ == "__main__":
    main()